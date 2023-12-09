import os
import sys
import openpyxl
from openpyxl.styles import PatternFill

# Load the Excel workbook
workbook = openpyxl.load_workbook(sys.argv[1])
sheet = workbook.active

# # Deleting rows (for example, rows 3 and 5)
# rows_to_delete = [1,]
# for row_idx in reversed(rows_to_delete):
#     sheet.delete_rows(row_idx)

# Hiding columns (for example, columns B and D)
columns_to_hide = ['A', 'B', 'C', 'D', 'E', 'G', 'H', 'J', 'K']
for col in columns_to_hide:
    sheet.column_dimensions[col].hidden = True

# Color-coding rows based on conditions
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
yellow_fill = PatternFill(start_color="FFFF32", end_color="FFFF32", fill_type="solid")

for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=14):
    condition_value = row[11].value  # Assuming the condition value is in the third column
    # print(condition_value)
    if condition_value == "ZTA":
        for cell in row:
            cell.fill = green_fill
    elif condition_value == "AGI":
        for cell in row:
            cell.fill = yellow_fill

# Save the modified workbook
base_filename, file_extension = os.path.splitext(sys.argv[1])
modified_filename = base_filename + "_colored" + file_extension
workbook.save(modified_filename)